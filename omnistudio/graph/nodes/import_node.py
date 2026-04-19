import csv
import openpyxl
import os
import re
from typing import Dict, List
from graph.state import WorkflowState


class ImportError(Exception):
    """Erreur d'import avec message utilisateur."""
    pass


FORMAT_ERRORS = {
    "csv":  "Le fichier CSV doit contenir au moins 2 colonnes (identifiant et texte), séparées par virgule, point-virgule ou tabulation.",
    "txt":  "Le fichier texte est vide ou ne contient aucun paragraphe (paragraphes séparés par une ligne vide).",
    "docx": "Le fichier Word ne contient aucun paragraphe de texte exploitable. Les titres, tableaux et images sont ignorés.",
    "pdf":  "Ce PDF ne contient pas de texte extractible. Les PDF scannés (images) ne sont pas supportés.",
    "xlsx": "Le fichier Excel doit contenir au moins 2 colonnes (identifiant en A, texte en B).",
    "md":   "Aucune étape trouvée. Le fichier Markdown doit utiliser le format : ### Étape 1 suivi du texte.",
}


def _make_step(step_id: str, text: str) -> Dict:
    """Crée un dict étape avec tous les champs requis."""
    return {
        "step_id": step_id,
        "text_original": text,
        "text_tts": "",
        "cleaning_status": "pending",
        "language_override": None,
        "speed_factor": 1.0,
    }


def _read_text_file(filepath: str) -> str:
    """Lit un fichier texte avec fallback encodage UTF-8 -> Latin-1."""
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            return f.read()
    except UnicodeDecodeError:
        with open(filepath, "r", encoding="latin-1") as f:
            return f.read()


def _split_long_paragraphs(paragraphs: List[str], max_words: int = 100) -> List[str]:
    """Découpe les paragraphes > max_words mots en phrases."""
    result = []
    for p in paragraphs:
        if len(p.split()) <= max_words:
            result.append(p)
        else:
            sentences = re.split(r'(?<=[.!?])\s+', p)
            chunk = ""
            for s in sentences:
                if chunk and len((chunk + " " + s).split()) > max_words:
                    result.append(chunk.strip())
                    chunk = s
                else:
                    chunk = (chunk + " " + s).strip() if chunk else s
            if chunk:
                result.append(chunk.strip())
    return result


def import_scenario(state: WorkflowState) -> Dict:
    """Nœud d'import : lit le fichier source et génère la liste des étapes."""
    source_file = state.get("source_file")
    if not source_file or not os.path.exists(source_file):
        raise ImportError("Fichier introuvable.")

    ext = os.path.splitext(source_file)[1].lower()
    try:
        if ext == ".xlsx":
            new_steps = _parse_xlsx(source_file, state.get("excel_sheet", "PLAN"))
        elif ext == ".md":
            new_steps = _parse_markdown(source_file)
        elif ext == ".csv":
            new_steps = _parse_csv(source_file)
        elif ext == ".txt":
            new_steps = _parse_txt(source_file)
        elif ext == ".docx":
            new_steps = _parse_docx(source_file)
        elif ext == ".pdf":
            new_steps = _parse_pdf(source_file)
        else:
            raise ImportError(f"Format {ext} non supporté.")
    except ImportError:
        raise
    except Exception as e:
        fmt = ext.lstrip(".")
        error_msg = FORMAT_ERRORS.get(fmt, f"Erreur lors du parsing ({e}).")
        # Log l'exception originale pour debugging
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Erreur parsing {fmt}: {e}", exc_info=True)
        raise ImportError(error_msg)

    if not new_steps:
        fmt = ext.lstrip(".")
        raise ImportError(FORMAT_ERRORS.get(fmt, "Aucune étape trouvée dans ce fichier."))

    # Gestion de l'import incrémental
    existing_steps = state.get("steps") or []
    if not existing_steps:
        return {
            "steps": new_steps,
            "source_format": ext.lstrip("."),
            "iteration_count": 1,
        }

    existing_ids = {str(s["step_id"]) for s in existing_steps}
    for ns in new_steps:
        if str(ns["step_id"]) not in existing_ids:
            existing_steps.append(ns)

    return {"steps": existing_steps, "iteration_count": 1}


def _parse_xlsx(filepath: str, sheet_name: str = "PLAN") -> List[Dict]:
    """Parse un fichier Excel (col A = step_id, col B = texte)."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]
    steps = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 2:
            continue
        step_id = str(row[0]).strip() if row[0] is not None else None
        text = str(row[1]).strip() if row[1] is not None else None
        if step_id and text:
            steps.append(_make_step(step_id, text))
    wb.close()
    return steps


def _parse_markdown(filepath: str) -> List[Dict]:
    """Parse un fichier Markdown (blocs ### Étape N)."""
    with open(filepath, "r", encoding="utf-8") as f:
        content = f.read()

    pattern = r"###\s+[ÉE]tape\s+(\d+)\s*\n+(.*?)(?=\n+###\s+[ÉE]tape|\Z)"
    matches = re.findall(pattern, content, re.DOTALL | re.IGNORECASE)

    steps = []
    for step_id, text in matches:
        text = text.strip()
        if text:
            steps.append(_make_step(step_id, text))
    return steps


def _parse_csv(filepath: str) -> List[Dict]:
    """Parse un fichier CSV (col 1 = step_id, col 2 = texte)."""
    raw = _read_text_file(filepath)
    try:
        dialect = csv.Sniffer().sniff(raw[:2048], delimiters=';,\t')
    except csv.Error:
        dialect = csv.excel
    # Vérification : si le Sniffer a choisi ',' mais que la 1re ligne de données
    # contient un ';' et pas de ',' entre les 2 premiers champs, c'est probablement ';'
    lines = raw.strip().splitlines()
    if len(lines) >= 2 and dialect.delimiter == ',':
        header = lines[0]
        if ';' in header and header.count(';') >= 1:
            dialect.delimiter = ';'
    reader = csv.reader(raw.splitlines(), dialect)
    next(reader, None)  # skip header
    steps = []
    for row in reader:
        if len(row) < 2:
            continue
        step_id = row[0].strip()
        text = row[1].strip()
        if step_id and text:
            steps.append(_make_step(step_id, text))
    return steps


def _parse_txt(filepath: str) -> List[Dict]:
    """Parse un fichier texte brut (paragraphes séparés par ligne vide)."""
    content = _read_text_file(filepath)
    paragraphs = [p.strip() for p in re.split(r"\n\s*\n", content) if p.strip()]
    paragraphs = _split_long_paragraphs(paragraphs)
    return [_make_step(str(i + 1), text) for i, text in enumerate(paragraphs)]


def _parse_docx(filepath: str) -> List[Dict]:
    """Parse un fichier Word (.docx) — paragraphes de corps uniquement."""
    from docx import Document
    doc = Document(filepath)
    paragraphs = []
    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue
        if para.style and para.style.name and para.style.name.startswith("Heading"):
            continue
        paragraphs.append(text)
    paragraphs = _split_long_paragraphs(paragraphs)
    if not paragraphs:
        return []
    return [_make_step(str(i + 1), text) for i, text in enumerate(paragraphs)]


def _parse_pdf(filepath: str) -> List[Dict]:
    """Parse un fichier PDF — extraction texte page par page."""
    import pdfplumber
    full_text = ""
    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                full_text += text + "\n\n"
    if len(full_text.strip()) < 20:
        return []
    paragraphs = [p.strip() for p in re.split(r"\n\s*\n", full_text) if p.strip()]
    paragraphs = _split_long_paragraphs(paragraphs)
    return [_make_step(str(i + 1), text) for i, text in enumerate(paragraphs)]
